from openpyxl import Workbook, load_workbook
import time
wb = None
ws = None

"""This method returns the cell of the searched value"""
def search_spesific_value(searched_value):
    global wb
    global ws
    for row in ws:
        for cell in row:
            if cell.value == searched_value:
                return cell
"""This method returns index number of the given cell"""
def get_index(cell):
    return cell.row,cell.column
"""This method will connect to the file in the given path"""
def connect(path):
    global wb
    wb = load_workbook(path, data_only=True)
"""This method will navigate to the sheet in the connected file"""
def navigate_sheet(sheet=None):
    global ws
    if sheet:
        ws = wb[sheet]
    else:
        ws=wb.active
def get_max_row():
    global ws
    return  ws.max_row
def get_max_column():
    global ws
    return  ws.max_column
def get_data(row,col):
    global ws
    return ws.cell(row=row, column=col)